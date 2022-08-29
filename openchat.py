import csv
from openpyxl import load_workbook, Workbook

f = open('kakao.csv', 'r')
kakao_csv = csv.reader(f)
load_wb = load_workbook("./kakao_openchat_fashion_startsup.xlsx", data_only=True)
load_ws1 = load_wb['Sheet1']
load_ws2 = load_wb['Sheet2']
write_wb = Workbook()
write_wb_1 = write_wb.create_sheet("Sheet1")
write_wb_2 = write_wb.create_sheet("Sheet2")

words = {}

for line in kakao_csv:
    if line[1] == "방장봇":
        continue
    if line[2].split(".")[0].split(" ")[-1] == "들어왔습니다":
        continue
    if line[2].split(" ")[-1] == "나갔습니다.":
        continue
    for word in line[2].split(" "):
        if word in words:
            words[word] += 1
        else:
            words[word] = 0
    write_wb_1.append(line)

for text in sorted(words.items(), key=lambda item: item[1], reverse=True):
    if(text[0] == " "):
        continue
    write_wb_2.append([text[0], text[1]])

write_wb.save('./kakao_openchat_fashion_startsup.xlsx')

f.close()
